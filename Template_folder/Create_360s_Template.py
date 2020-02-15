from openpyxl import *
from openpyxl.styles import Font, Alignment
import os

#Create an array listing all files in the directory
file_arr = os.listdir()

#If the template is already in the directory, remove it.
if '360_Reviews_Template.xlsx' in file_arr:
    os.remove('360_Reviews_Template.xlsx')

#Load the Employees and Questions files
eb = load_workbook('Employees.xlsx')
es = eb.active
qb = load_workbook('Questions.xlsx')

#Create a dictionary of position types (keys) and lists of employees (values)
employee_dict = {}
for col in es.iter_cols(max_col=4):
    l = []
    for cell in col:
        if cell.value != None:
            l.append(cell.value)
    m = l[1:]
    m.sort()
    employee_dict[l[0]] = m
    
#Create a list of questions from the master list.
questions = []
for row in qb.active:
    for cell in row:
        questions.append(cell.value)

#Format the template
file_name = "360_Reviews_Template.xlsx"
file = Workbook()
file.active.title = "Dispatchers"
file.create_sheet("Supervisors")
file.create_sheet("Admin")
file.create_sheet("Director")

for sheet in file.sheetnames:
    #Set column width. This assumes there are only 3 questions being asked to
    #make the process simplified.
    file[sheet].column_dimensions['A'].width = 25
    file[sheet].column_dimensions['B'].width = 25
    file[sheet].column_dimensions['C'].width = 25
    file[sheet].column_dimensions['D'].width = 25
    
    #Header formatting
    file[sheet].merge_cells('A1:D1')
    file[sheet]['A1'] = "Company ABC" # Change company name here
    file[sheet]['A1'].font = Font(bold=True)
    file[sheet]['A1'].alignment = Alignment(horizontal='center')
    
    file[sheet].merge_cells('A2:D2')
    file[sheet]['A2'] = "360 Reviews"
    file[sheet]['A2'].font = Font(bold=True)
    file[sheet]['A2'].alignment = Alignment(horizontal='center')

    file[sheet]['A3'] = "Reviewer:"
    file[sheet]['A3'].font = Font(bold=True)
    file[sheet]['A3'].alignment = Alignment(horizontal='right')

    file[sheet]['A4'] = "Employee"
    file[sheet]['A4'].font = Font(bold=True)

    file[sheet]['C3'] = "Date:"
    file[sheet]['C3'].font = Font(bold=True)
    file[sheet]['C3'].alignment = Alignment(horizontal='right')
    
    #Add employee names to their relevant sheets.
    employees = employee_dict[sheet]
    for x in range(len(questions)):
        file[sheet].cell(row = 4, column = x+2, value = questions[x]).font = Font(bold=True)
        for y in range(len(employees)):
            file[sheet].cell(row = y+5, column = 1, value = employees[y])

    #Wrap text for all question and answer cells.
    for rows in file[sheet].iter_rows(min_row = 4, min_col = 2):
        for cell in rows:
            cell.alignment = Alignment(wrap_text=True)

file.save(file_name)    