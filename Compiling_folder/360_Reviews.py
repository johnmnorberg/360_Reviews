from openpyxl import *
from openpyxl.styles import Font, Alignment
import os
import datetime
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

#Define suggested hierarchy
dispatch_hierarchy = "1, the supervisors email"
supervisor_hierarchy = "2, the operations manager and the director"
admin_hierarchy = "1, the director"
director_hierarchy = "1, the chair of the executive board"

#Function that creates lists of email addresses for various employee types.
def get_email(employee_type, employee_hierarchy):
    
    email_list = []
    
    #Establish the number of email addresses that will be accepted.
    while True:        
        try:
            count = int(input("How many email addresses are you sending '{}.xlsx' to? (Suggested: {}) ".format(employee_type, employee_hierarchy)))
        except ValueError:
            print("That is not a valid answer.")
            continue
        else:
            break
        
    #Obtain email addresses
    for x in range(count):
        email_list.append(input("Email #{}: Sending '{}.xlsx' to... ".format(x+1, employee_type)))

    return email_list

#Lists of email addresses. Dispatchers.xlsx gets sent to supervisors, Supervisors.xlsx
#gets sent to the ops manager, Admin.xlsx gets sent to the director, and Director.xlsx
#gets sent to the executive board.
dispatch_email = get_email("Dispatchers", dispatch_hierarchy)
supervisor_email = get_email("Supervisors", supervisor_hierarchy)
admin_email = get_email("Admin", admin_hierarchy)
director_email = get_email("Director", director_hierarchy)

#Get today's date in MM/DD/YY format
date = datetime.datetime.now()
date = date.strftime("%x")

#Create an array listing all files in the directory
file_arr = os.listdir()

#Load the Employees and Question files
eb = load_workbook('Employees.xlsx')
es = eb.active
qb = load_workbook('Questions.xlsx')
  
#Create a dictionary of position types (keys) and lists of employees (values)
employee_dict = {}
for col in es.iter_cols(max_col=4):
    l = []
    for cell in col:
        if cell.value != None:
            l.append(cell.value)
    m = l[1:]
    m.sort()
    employee_dict[l[0]] = m
 
#Create a list of all employees
employees = []
for value_list in employee_dict.values():
    for name in value_list:
        employees.append(name)
employees.sort()

#Create blank employee files if they were not submitted. This is similar
#to the template, but with a note that it was not completed. This will prevent
#the application from failing due to a missing (expected) file.
#
#IMPORTANT - IF THERE IS A MISSING FILE, THE SCRIPT WILL FAIL ON THE FIRST ATTEMPT
#SIMPLY RE-RUN AGAIN AS IT NEEDS THE FIRST ATTEMPT TO CREATE MISSING FILES.
for name in employees:
    if name + '.xlsx' not in file_arr:
        file_name = name + '.xlsx'
        file = Workbook()
        file.active.title = "Dispatchers"
        file.create_sheet("Supervisors")
        file.create_sheet("Admin")
        file.create_sheet("Director")

        for sheet in file.sheetnames:
            file[sheet].column_dimensions['A'].width = 25
            file[sheet].column_dimensions['B'].width = 25
            file[sheet].column_dimensions['C'].width = 25
            file[sheet].column_dimensions['D'].width = 25
            
            file[sheet].merge_cells('A1:D1')
            file[sheet]['A1'] = "Company ABC"
            file[sheet]['A1'].font = Font(bold=True)
            file[sheet]['A1'].alignment = Alignment(horizontal='center')
            
            file[sheet].merge_cells('A2:D2')
            file[sheet]['A2'] = "360 Reviews"
            file[sheet]['A2'].font = Font(bold=True)
            file[sheet]['A2'].alignment = Alignment(horizontal='center')
        
            file[sheet]['A3'] = "Reviewer:"
            file[sheet]['A3'].font = Font(bold=True)
            file[sheet]['A3'].alignment = Alignment(horizontal='right')
            
            file[sheet]['B3'] = name
        
            file[sheet]['C3'] = "Date:"
            file[sheet]['C3'].font = Font(bold=True)
            file[sheet]['C3'].alignment = Alignment(horizontal='right')
            
            file[sheet]['D3'] = "DID NOT COMPLETE"
            file[sheet]['D3'].font = Font(color="ff0000")
            
            file.save(file_name)    
            
#Create a list of questions from the master list.
questions = []
for row in qb.active:
    for cell in row:
        questions.append(cell.value)
        
#Remove unncessary files from the file array
file_arr.remove('Employees.xlsx')
file_arr.remove('Questions.xlsx')
if '360_Reviews.py' in file_arr:
	file_arr.remove('360_Reviews.py')
if '360_Reviews.exe' in file_arr:
	file_arr.remove('360_Reviews.exe')

#Create a function that loads workbooks
def load_create_wb(file_name, employee_type):
    #If the output file already exists, load it and remove it from the file array.
    if file_name in file_arr:
        wb = load_workbook(file_name)
        file_arr.remove(file_name)
    
    #If the output file does not exist, create it.
    else:
        wb = Workbook()
        
        #Extract the list of employees for the specified employee type.
        l = employee_dict[employee_type]
        
        #Name the first sheet after the first employee
        wb.active.title = l[0]
        
        #Create sheets for each additional employee and name the sheet after the employee.
        for x in range(len(l)-1):
            wb.create_sheet(l[x+1])
            
        #Set column width. Add the headers and employee names to the output file.
        for sheet in wb.sheetnames:
            wb[sheet].column_dimensions['A'].width = 25
            wb[sheet].column_dimensions['B'].width = 25
            wb[sheet].column_dimensions['C'].width = 25
            wb[sheet].column_dimensions['D'].width = 25
            wb[sheet].column_dimensions['E'].width = 25
            
            wb[sheet].merge_cells('A1:D1')
            wb[sheet]['A1'] = "Company ABC"
            wb[sheet]['A1'].font = Font(bold=True)
            wb[sheet]['A1'].alignment = Alignment(horizontal='center')
          
            wb[sheet].merge_cells('A2:D2')
            wb[sheet]['A2'] = "360 Reviews"
            wb[sheet]['A2'].font = Font(bold=True)
            wb[sheet]['A2'].alignment = Alignment(horizontal='center')
            
            wb[sheet]['B3'] = "Reviews for: " + sheet
            wb[sheet]['B3'].font = Font(bold=True)
            
            wb[sheet]['D3'] = "Date: " + date
            wb[sheet]['D3'].font = Font(bold=True)
            
            wb[sheet]['A4'] = "Reviewer"
            wb[sheet]['A4'].font = Font(bold=True)
                
            for x in range(len(questions)):
                wb[sheet].cell(row = 4, column = x+2, value = questions[x]).font = Font(bold=True)
                for y in range(len(employees)):
                    wb[sheet].cell(row = y+5, column = 1, value = employees[y])
                    
            #Wrap text for all question and answer cells.
            for rows in wb[sheet].iter_rows(min_row = 4, min_col = 2):
                for cell in rows:
                    cell.alignment = Alignment(wrap_text=True)
                            
    return wb

#Create workbooks for dispatchers, supervisors, and admin
tb = load_create_wb('Dispatchers.xlsx', 'Dispatchers')
sb = load_create_wb('Supervisors.xlsx', 'Supervisors')
ab = load_create_wb('Admin.xlsx', 'Admin')
db = load_create_wb('Director.xlsx', 'Director')

#Create a list containing each input file.
file_inputs = []
for x in file_arr:
    file_inputs.append(load_workbook(x)) 

#Create one list of answers for each employee type. The list loops through each submission.
#That is, the first row from each submission will be recorded, then the second row, and so on.
#Each elemet will be a cell from the input file.   
def answers(employee_type):
    l = []
    for file in file_inputs:
        for row in file[employee_type].iter_rows(min_col = 2, max_col = len(questions)+1, min_row = 5, max_row = len(employees)+4):
            for cell in row:
                l.append(cell.value)
    return l

dispatch_answers = answers('Dispatchers')      
supervisor_answers = answers('Supervisors')
admin_answers = answers('Admin')
director_answers = answers('Director')

#Create a list of sublists where each sublist is a row from each input file.
#The first element will be the first row from the first submission. The second element will be the first row from the second submission.
#Once all of the first rows are accounted for, it will loop to the second row, and so on.
def answers_by_row(employee_answers):
    l = []
    for i in range(0, len(employee_answers), len(questions)):
        sublist = employee_answers[i:i+len(questions)]
        l.append(sublist)
    return l

dispatch_answers_by_row = answers_by_row(dispatch_answers)
supervisor_answers_by_row = answers_by_row(supervisor_answers)
admin_answers_by_row = answers_by_row(admin_answers)
director_answers_by_row = answers_by_row(director_answers)
        
#Take the previous lists by rows and turn them into an array. Each element of this list will be the array of answers for each sheet.
#That is the first element will be list of sublists where each sublist is the output rows for the first sheet.
#The second element will be a list of sublists where each sublist is the output rows for the second sheet, and so on.
def answers_by_cell(employee_answers_by_row):
    l = []
    for i in range(0, len(employee_answers_by_row), len(employees)):
        sublist = employee_answers_by_row[i:i+len(employees)]
        l.append(sublist)
    return l

dispatch_answers_by_cell = answers_by_cell(dispatch_answers_by_row)
supervisor_answers_by_cell = answers_by_cell(supervisor_answers_by_row)
admin_answers_by_cell = answers_by_cell(admin_answers_by_row)
director_answers_by_cell = answers_by_cell(director_answers_by_row)

#Take the previous lists/matrices and "project" them onto their corresponding sheets.
def compile_wb(employee_type, wb, employee_answers_by_cell):
    l = employee_dict[employee_type]    
    for x in range(len(employees)):
        for y in range(len(l)):
            for z in range(len(questions)):
                wb[l[y]].cell(row=x+5, column=z+2, value=employee_answers_by_cell[x][y][z])
    return wb

compile_wb('Dispatchers', tb, dispatch_answers_by_cell)
compile_wb('Supervisors', sb, supervisor_answers_by_cell)
compile_wb('Admin', ab, admin_answers_by_cell)
compile_wb('Director', db, director_answers_by_cell)

tb.save('Dispatchers.xlsx')
sb.save('Supervisors.xlsx')
ab.save('Admin.xlsx')   
db.save('Director.xlsx')


#Email portion
def send_email(email_list, employee_type, filename):

    body = """Hello,\n\nAttached are the 360 reviews for Company ABC {}. Please do not reply to this automated message.""".format(employee_type)
    
    # Create a multipart message and set headers
    message = MIMEMultipart()
    message["From"] = "360@companyabc.org" #Insert sender email address
    message["To"] = ", ".join(email_list)
    message["Subject"] = "Company ABC: 360 Reviews"

    # Add body to email
    message.attach(MIMEText(body, "plain"))

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(part)

    # Add header as key/value pair to attachment part
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {filename}",
    )

    # Add attachment to message and convert message to string
    message.attach(part)
    text = message.as_string()
    
    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    with smtplib.SMTP("my_server_address", 25) as server: #Insert company server address
        server.sendmail("360@companyabc.org", ", ".join(email_list), text)
    
if len(dispatch_email) > 0:
    send_email(dispatch_email, "dispatchers", "Dispatchers.xlsx")
if len(supervisor_email) > 0:
    send_email(supervisor_email, "supervisors", "Supervisors.xlsx")
if len(admin_email) > 0:
    send_email(admin_email, "admin", "Admin.xlsx")
if len(director_email) > 0:
    send_email(director_email, "director", "Director.xlsx")
    
#Delete the output files after emailing them
os.remove("Dispatchers.xlsx")
os.remove("Supervisors.xlsx")
os.remove("Admin.xlsx")
os.remove("Director.xlsx")

input("Press ENTER to close.")